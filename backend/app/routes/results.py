import os
import re
import csv
import io
import json
import logging
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from bson import ObjectId

from ..config.database import db
from ..routes.auth import get_current_admin
from ..utils.storage import get_file_url, save_uploaded_file, delete_file

logger = logging.getLogger(__name__)
router = APIRouter()

# =====================================================
# PYDANTIC SCHEMAS
# =====================================================

class SubjectScore(BaseModel):
    subject: str
    max_marks: float = 100.0
    obtained_marks: float
    grade: Optional[str] = ""
    remarks: Optional[str] = ""

class StudentResultSchema(BaseModel):
    student_name: str
    admission_no: str
    roll_no: str
    father_name: str
    class_name: str = Field(..., alias="class")
    section: str
    academic_year: str
    subjects: List[SubjectScore]

    class Config:
        allow_population_by_field_name = True

# =====================================================
# HELPERS FOR CALCULATION
# =====================================================

def calculate_result_metrics(subjects: List[dict]):
    total_obtained = sum(float(s.get("obtained_marks", 0)) for s in subjects)
    total_max = sum(float(s.get("max_marks", 100)) for s in subjects)
    
    percentage = (total_obtained / total_max * 100) if total_max > 0 else 0.0
    
    # KVS style grading: A+ (>=90), A (>=80), B (>=70), C (>=60), D (>=50), E (>=35), F (<35)
    if percentage >= 90:
        grade = "A+"
    elif percentage >= 80:
        grade = "A"
    elif percentage >= 70:
        grade = "B"
    elif percentage >= 60:
        grade = "C"
    elif percentage >= 50:
        grade = "D"
    elif percentage >= 35:
        grade = "E"
    else:
        grade = "F"
        
    # Pass/Fail calculation (Fail if any subject has < 35% marks)
    status = "Pass"
    for s in subjects:
        sub_pct = (float(s.get("obtained_marks", 0)) / float(s.get("max_marks", 100)) * 100) if float(s.get("max_marks", 100)) > 0 else 0.0
        if sub_pct < 35.0:
            status = "Fail"
            break
            
    # Division calculation
    if status == "Fail":
        division = "Fail"
    elif percentage >= 60.0:
        division = "First"
    elif percentage >= 45.0:
        division = "Second"
    elif percentage >= 35.0:
        division = "Third"
    else:
        division = "Fail"
        
    return {
        "total": total_obtained,
        "percentage": round(percentage, 2),
        "division": division,
        "grade": grade,
        "status": status
    }

def helper_serialize_id(doc):
    if doc and "_id" in doc:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
    return doc

async def recalculate_class_toppers(class_name: str):
    try:
        cursor = db.results.find({"class": class_name.strip()}).sort([("percentage", -1)])
        class_results = await cursor.to_list(100)
        
        if class_results:
            # Preserve existing photos for toppers of this class
            existing_photos = {}
            async for t in db.result_toppers.find({"class": class_name}):
                name = t.get("student_name", "").strip()
                if name:
                    existing_photos[name] = t.get("photo_url", "")
            
            # Delete old toppers for this class
            await db.result_toppers.delete_many({"class": class_name})
            
            top_n = min(3, len(class_results))
            for i in range(top_n):
                student = class_results[i]
                rank_str = "1st" if i == 0 else "2nd" if i == 1 else "3rd"
                medal_str = "Gold Medalist" if i == 0 else "Silver Medalist" if i == 1 else "Bronze Medalist"
                
                total_max = sum(float(s.get("max_marks", 100)) for s in student.get("subjects", []))
                
                total_val = student.get('total', 0)
                if isinstance(total_val, float) and total_val.is_integer():
                    total_val = int(total_val)
                    
                if total_max > 0:
                    if isinstance(total_max, float) and total_max.is_integer():
                        total_max = int(total_max)
                    marks_str = f"{total_val} / {total_max}"
                else:
                    marks_str = f"{total_val}"
                
                s_name = student.get("student_name", "").strip()
                photo_url = existing_photos.get(s_name, "")
                
                topper_doc = {
                    "student_name": s_name,
                    "class": class_name,
                    "marks": marks_str,
                    "percentage": float(student.get("percentage", 0.0)),
                    "rank": rank_str,
                    "medal": medal_str,
                    "photo_url": photo_url,
                    "created_at": datetime.utcnow()
                }
                await db.result_toppers.insert_one(topper_doc)
            logger.info(f"Auto-recalculated top {top_n} performers for Class {class_name}")
    except Exception as e:
        logger.error(f"Failed to auto-recalculate toppers: {str(e)}")

# =====================================================
# PUBLIC PORTAL ENDPOINTS
# =====================================================

@router.get("/results")
async def get_public_portal_data():
    """Retrieve all public data for result.html main portal."""
    hero = await db.site_settings.find_one({"key": "results_hero"}) or {
        "title": "Academic Results",
        "subtitle": "View Academic Performance and Examination Results for Classes VI to XII",
        "banner_url": ""
    }
    hero = helper_serialize_id(hero)
    
    stats = await db.result_statistics.find_one({"key": "global_statistics"}) or {
        "pass_percentage": "98%",
        "distinctions": "120",
        "first_division": "210",
        "students_appeared": "350",
        "students_passed": "343",
        "topper_marks": "98.8%"
    }
    stats = helper_serialize_id(stats)
    
    notices = await db.result_notices.find().sort([("pinned", -1), ("notice_date", -1)]).to_list(100)
    for n in notices:
        helper_serialize_id(n)
        
    # ONLY class 10 ("X") and class 12 ("XII") toppers
    toppers = await db.result_toppers.find({"class": {"$in": ["X", "XII"]}}).sort([("percentage", -1)]).to_list(100)
    for t in toppers:
        helper_serialize_id(t)
        
    # Get general class summary stats for cards
    class_cards = []
    classes = ["VI", "VII", "VIII", "IX", "X", "XI", "XII"]
    for c in classes:
        # Calculate summary numbers dynamically from actual results database
        cursor = db.results.find({"class": c})
        all_results = await cursor.to_list(1000)
        
        appeared = len(all_results)
        passed = sum(1 for r in all_results if r.get("status") == "Pass")
        pass_pct = f"{round((passed / appeared * 100), 1)}%" if appeared > 0 else "0%"
        
        top_score = "N/A"
        if all_results:
            max_pct = max(float(r.get("percentage", 0)) for r in all_results)
            top_score = f"{max_pct}%"
            
        class_cards.append({
            "class_name": f"Class {c}",
            "class_code": c,
            "appeared": appeared,
            "passed": passed,
            "pass_percentage": pass_pct,
            "top_score": top_score
        })
        
    return {
        "hero": hero,
        "statistics": stats,
        "notices": notices,
        "toppers": toppers,
        "class_cards": class_cards
    }


CLASS_KEY_MAP = {
    "class6": "VI",
    "class7": "VII",
    "class8": "VIII",
    "class9": "IX",
    "class10": "X",
    "class11": "XI",
    "class12": "XII"
}

@router.get("/results/public/{class_key}")
async def get_public_class_results(class_key: str):
    """Serve public results page for a specific class portal."""
    normalized_key = class_key.strip().lower().replace("-", "")
    class_name = CLASS_KEY_MAP.get(normalized_key, class_key.upper())
    
    # Retrieve all student results for this class sorted by percentage descending
    cursor = db.results.find({"class": class_name}).sort([("percentage", -1)])
    results = await cursor.to_list(1000)
    
    # Retrieve toppers for this class from result_toppers collection
    toppers_cursor = db.result_toppers.find({"class": class_name}).sort([("percentage", -1)])
    toppers = await toppers_cursor.to_list(100)
    
    # Map toppers to structure expected by results-public.js: {name, rank, marks}
    toppers_mapped = []
    for t in toppers:
        toppers_mapped.append({
            "name": t.get("student_name", ""),
            "rank": t.get("rank", ""),
            "marks": t.get("marks", "")
        })
        
    # Map merit list (all students sorted) to: {name, rank, marks}
    merit_mapped = []
    for idx, r in enumerate(results):
        merit_mapped.append({
            "name": r.get("student_name", ""),
            "rank": str(idx + 1),
            "marks": f"{r.get('percentage', 0.0)}%"
        })
        
    # Map all students list to structure expected by results-public.js: {name, roll_no, dob, marks, percentage}
    students_mapped = []
    for r in results:
        students_mapped.append({
            "name": r.get("student_name", ""),
            "roll_no": r.get("roll_no", ""),
            "dob": "-",  # DOB is not stored in the schema, using placeholder
            "marks": r.get("total", 0.0),
            "percentage": f"{r.get('percentage', 0.0)}%"
        })
        
    return {
        "title": f"Class {class_name} Board Results",
        "subtitle": "Academic Session Result",
        "toppers": toppers_mapped,
        "merit_list": merit_mapped,
        "students": students_mapped
    }

@router.get("/results/public/{class_key}/student")
async def get_public_individual_student_result(class_key: str, roll_no: str = Query(...), dob: Optional[str] = Query(None)):
    """Retrieve an individual student's result."""
    normalized_key = class_key.strip().lower().replace("-", "")
    class_name = CLASS_KEY_MAP.get(normalized_key, class_key.upper())
    
    student = await db.results.find_one({"roll_no": roll_no.strip(), "class": class_name})
    if not student:
        raise HTTPException(status_code=404, detail="No student result found for the provided Roll Number.")
        
    return {
        "student": {
            "name": student.get("student_name", ""),
            "roll_no": student.get("roll_no", ""),
            "class_name": f"Class {student.get('class', '')}",
            "dob": dob or "-",
            "marks": student.get("total", 0.0),
            "percentage": f"{student.get('percentage', 0.0)}%"
        },
        "class_name": f"Class {student.get('class', '')}"
    }

# =====================================================
# HERO CONFIG ENDPOINTS
# =====================================================

@router.get("/results/hero")
async def get_results_hero():
    hero = await db.site_settings.find_one({"key": "results_hero"}) or {
        "title": "Academic Results",
        "subtitle": "View Academic Performance and Examination Results for Classes VI to XII",
        "banner_url": ""
    }
    return helper_serialize_id(hero)

@router.put("/results/hero")
async def update_results_hero(
    title: str = Form(...),
    subtitle: str = Form(...),
    banner_file: Optional[UploadFile] = File(None),
    admin=Depends(get_current_admin)
):
    current = await db.site_settings.find_one({"key": "results_hero"}) or {}
    banner_url = current.get("banner_url", "")
    
    if banner_file:
        file_id = await save_uploaded_file(db, banner_file, category="results")
        banner_url = get_file_url(file_id)
        
    doc = {
        "key": "results_hero",
        "title": title.strip(),
        "subtitle": subtitle.strip(),
        "banner_url": banner_url,
        "updated_at": datetime.utcnow()
    }
    
    await db.site_settings.update_one(
        {"key": "results_hero"},
        {"$set": doc},
        upsert=True
    )
    return {"message": "Hero settings updated successfully"}

# =====================================================
# ACADEMIC STATISTICS ENDPOINTS
# =====================================================

@router.get("/results/statistics")
async def get_results_statistics():
    stats = await db.result_statistics.find_one({"key": "global_statistics"}) or {
        "pass_percentage": "98%",
        "distinctions": "120",
        "first_division": "210",
        "students_appeared": "350",
        "students_passed": "343",
        "topper_marks": "98.8%"
    }
    return helper_serialize_id(stats)

@router.put("/results/statistics")
async def update_results_statistics(
    pass_percentage: str = Form(...),
    distinctions: str = Form(...),
    first_division: str = Form(...),
    students_appeared: str = Form(...),
    students_passed: str = Form(...),
    topper_marks: str = Form(...),
    admin=Depends(get_current_admin)
):
    doc = {
        "key": "global_statistics",
        "pass_percentage": pass_percentage.strip(),
        "distinctions": distinctions.strip(),
        "first_division": first_division.strip(),
        "students_appeared": students_appeared.strip(),
        "students_passed": students_passed.strip(),
        "topper_marks": topper_marks.strip(),
        "updated_at": datetime.utcnow()
    }
    
    await db.result_statistics.update_one(
        {"key": "global_statistics"},
        {"$set": doc},
        upsert=True
    )
    return {"message": "Academic statistics updated successfully"}

# =====================================================
# RESULT NOTICES ENDPOINTS
# =====================================================

@router.get("/results/notices")
async def get_results_notices():
    notices = await db.result_notices.find().sort([("pinned", -1), ("notice_date", -1)]).to_list(100)
    for n in notices:
        helper_serialize_id(n)
    return notices

@router.post("/results/notices")
async def create_result_notice(
    title: str = Form(...),
    notice_date: str = Form(...),
    content: str = Form(...),
    pinned: bool = Form(False),
    pdf_file: Optional[UploadFile] = File(None),
    admin=Depends(get_current_admin)
):
    pdf_url = ""
    pdf_name = ""
    if pdf_file:
        file_id = await save_uploaded_file(db, pdf_file, category="results")
        pdf_url = get_file_url(file_id)
        pdf_name = pdf_file.filename
        
    doc = {
        "title": title.strip(),
        "notice_date": notice_date.strip(),
        "content": content.strip(),
        "pinned": pinned,
        "pdf_url": pdf_url,
        "pdf_name": pdf_name,
        "created_at": datetime.utcnow()
    }
    
    res = await db.result_notices.insert_one(doc)
    return {"message": "Notice created successfully", "id": str(res.inserted_id)}

@router.put("/results/notices/{id}")
async def update_result_notice(
    id: str,
    title: str = Form(...),
    notice_date: str = Form(...),
    content: str = Form(...),
    pinned: bool = Form(False),
    pdf_file: Optional[UploadFile] = File(None),
    remove_pdf: bool = Form(False),
    admin=Depends(get_current_admin)
):
    current = await db.result_notices.find_one({"_id": ObjectId(id)})
    if not current:
        raise HTTPException(status_code=404, detail="Notice not found")
        
    pdf_url = current.get("pdf_url", "")
    pdf_name = current.get("pdf_name", "")
    
    if remove_pdf:
        pdf_url = ""
        pdf_name = ""
        
    if pdf_file:
        file_id = await save_uploaded_file(db, pdf_file, category="results")
        pdf_url = get_file_url(file_id)
        pdf_name = pdf_file.filename
        
    doc = {
        "title": title.strip(),
        "notice_date": notice_date.strip(),
        "content": content.strip(),
        "pinned": pinned,
        "pdf_url": pdf_url,
        "pdf_name": pdf_name,
        "updated_at": datetime.utcnow()
    }
    
    await db.result_notices.update_one({"_id": ObjectId(id)}, {"$set": doc})
    return {"message": "Notice updated successfully"}

@router.delete("/results/notices/{id}")
async def delete_result_notice(id: str, admin=Depends(get_current_admin)):
    res = await db.result_notices.delete_one({"_id": ObjectId(id)})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notice not found")
    return {"message": "Notice deleted successfully"}

# =====================================================
# RESULT TOPPERS ENDPOINTS
# =====================================================

@router.get("/results/toppers")
async def get_results_toppers():
    toppers = await db.result_toppers.find().sort([("percentage", -1)]).to_list(100)
    for t in toppers:
        helper_serialize_id(t)
    return toppers

@router.post("/results/toppers")
async def create_result_topper(
    student_name: str = Form(...),
    class_name: str = Form(..., alias="class"),
    marks: str = Form(...),
    percentage: float = Form(...),
    rank: str = Form(...),
    medal: Optional[str] = Form(""),
    student_photo: Optional[UploadFile] = File(None),
    admin=Depends(get_current_admin)
):
    photo_url = ""
    if student_photo:
        file_id = await save_uploaded_file(db, student_photo, category="results")
        photo_url = get_file_url(file_id)
        
    doc = {
        "student_name": student_name.strip(),
        "class": class_name.strip(),
        "marks": marks.strip(),
        "percentage": percentage,
        "rank": rank.strip(),
        "medal": medal.strip(),
        "photo_url": photo_url,
        "created_at": datetime.utcnow()
    }
    
    res = await db.result_toppers.insert_one(doc)
    return {"message": "Topper profile created successfully", "id": str(res.inserted_id)}

@router.put("/results/toppers/{id}")
async def update_result_topper(
    id: str,
    student_name: str = Form(...),
    class_name: str = Form(..., alias="class"),
    marks: str = Form(...),
    percentage: float = Form(...),
    rank: str = Form(...),
    medal: Optional[str] = Form(""),
    student_photo: Optional[UploadFile] = File(None),
    remove_photo: bool = Form(False),
    admin=Depends(get_current_admin)
):
    current = await db.result_toppers.find_one({"_id": ObjectId(id)})
    if not current:
        raise HTTPException(status_code=404, detail="Topper not found")
        
    photo_url = current.get("photo_url", "")
    if remove_photo:
        photo_url = ""
        
    if student_photo:
        file_id = await save_uploaded_file(db, student_photo, category="results")
        photo_url = get_file_url(file_id)
        
    doc = {
        "student_name": student_name.strip(),
        "class": class_name.strip(),
        "marks": marks.strip(),
        "percentage": percentage,
        "rank": rank.strip(),
        "medal": medal.strip(),
        "photo_url": photo_url,
        "updated_at": datetime.utcnow()
    }
    
    await db.result_toppers.update_one({"_id": ObjectId(id)}, {"$set": doc})
    return {"message": "Topper profile updated successfully"}

@router.delete("/results/toppers/{id}")
async def delete_result_topper(id: str, admin=Depends(get_current_admin)):
    res = await db.result_toppers.delete_one({"_id": ObjectId(id)})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Topper not found")
    return {"message": "Topper profile deleted successfully"}

# =====================================================
# STUDENT RESULT CRUD ENDPOINTS
# =====================================================

@router.get("/results/class/{class_name}")
async def get_results_by_class(class_name: str):
    cursor = db.results.find({"class": class_name.strip()})
    results = await cursor.to_list(1000)
    for r in results:
        helper_serialize_id(r)
    return results

@router.get("/results/student/{admission_no}")
async def get_student_result(admission_no: str):
    result = await db.results.find_one({"admission_no": admission_no.strip()})
    if not result:
        raise HTTPException(status_code=404, detail="Student result not found")
    return helper_serialize_id(result)

@router.post("/results")
async def create_student_result(schema: StudentResultSchema, admin=Depends(get_current_admin)):
    # Check if student already exists in this academic year
    existing = await db.results.find_one({
        "admission_no": schema.admission_no.strip(),
        "academic_year": schema.academic_year.strip()
    })
    if existing:
        raise HTTPException(status_code=400, detail="Result already exists for this admission number and academic year")
        
    doc = schema.dict(by_alias=True)
    
    # Recalculate metrics defensively
    metrics = calculate_result_metrics(doc["subjects"])
    doc.update(metrics)
    
    doc["created_at"] = datetime.utcnow()
    doc["updated_at"] = datetime.utcnow()
    
    res = await db.results.insert_one(doc)
    # recalculate_class_toppers is now handled dynamically in front-end class portals
    return {"message": "Student result saved successfully", "id": str(res.inserted_id)}

@router.put("/results/{id}")
async def update_student_result(id: str, schema: StudentResultSchema, admin=Depends(get_current_admin)):
    current = await db.results.find_one({"_id": ObjectId(id)})
    if not current:
        raise HTTPException(status_code=404, detail="Result record not found")
        
    doc = schema.dict(by_alias=True)
    
    # Recalculate metrics dynamically
    metrics = calculate_result_metrics(doc["subjects"])
    doc.update(metrics)
    
    doc["updated_at"] = datetime.utcnow()
    
    await db.results.update_one({"_id": ObjectId(id)}, {"$set": doc})
    # recalculate_class_toppers is now handled dynamically in front-end class portals
    return {"message": "Student result updated successfully"}

@router.delete("/results/{id}")
async def delete_student_result(id: str, admin=Depends(get_current_admin)):
    current = await db.results.find_one({"_id": ObjectId(id)})
    if not current:
        raise HTTPException(status_code=404, detail="Result record not found")
    class_name = current.get("class", "")
    res = await db.results.delete_one({"_id": ObjectId(id)})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Result record not found")
    # recalculate_class_toppers is now handled dynamically in front-end class portals
    return {"message": "Result record deleted successfully"}

# =====================================================
# SEARCH ENDPOINT
# =====================================================

@router.get("/results/search")
async def search_student_results(
    admission_no: Optional[str] = Query(None),
    roll_no: Optional[str] = Query(None),
    student_name: Optional[str] = Query(None),
    class_name: Optional[str] = Query(None)
):
    query = {}
    
    if class_name:
        query["class"] = class_name.strip()
        
    if admission_no:
        query["admission_no"] = admission_no.strip()
    elif roll_no:
        query["roll_no"] = roll_no.strip()
    elif student_name:
        query["student_name"] = {"$regex": re.escape(student_name.strip()), "$options": "i"}
        
    if not query:
        raise HTTPException(status_code=400, detail="Search query parameters are required")
        
    cursor = db.results.find(query).sort([("percentage", -1)])
    results = await cursor.to_list(100)
    for r in results:
        helper_serialize_id(r)
    return results

# =====================================================
# EXCEL IMPORT/EXPORT
# =====================================================

@router.post("/results/import/{class_name}")
async def import_results(
    class_name: str,
    file: UploadFile = File(...),
    admin=Depends(get_current_admin)
):
    filename = (file.filename or "").strip()
    content = await file.read()
    
    rows = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    elif filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            
            headers = []
            for cell in sheet[1]:
                headers.append((cell.value or "").strip().lower())
                
            for r in range(2, sheet.max_row + 1):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                if not any(row_vals):
                    continue
                row_dict = {}
                for idx, h in enumerate(headers):
                    if h:
                        row_dict[h] = row_vals[idx]
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel library (openpyxl) is not installed on the server. Please upload a .csv file.")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Supported formats: .csv, .xlsx")
        
    imported = 0
    errors = []
    
    def find_row_val(row_dict, keywords):
        for k, v in row_dict.items():
            k_clean = str(k).lower().strip()
            for kw in keywords:
                if kw in k_clean:
                    return v
        return None

    # Process rows
    for index, r in enumerate(rows):
        try:
            student_name = str(find_row_val(r, ["student name", "student_name", "name"]) or "").strip()
            admission_no = str(find_row_val(r, ["admission no", "admission_number", "admission_no", "adm no", "admission"]) or "").strip()
            roll_no = str(find_row_val(r, ["roll no", "roll_number", "roll_no", "roll"]) or "").strip()
            father_name = str(find_row_val(r, ["father name", "father_name", "father's name", "father"]) or "").strip()
            section = str(find_row_val(r, ["section", "sec"]) or "A").strip()
            academic_year = str(find_row_val(r, ["academic year", "academic_year", "year"]) or "2025-2026").strip()
            
            if not student_name or not admission_no or not roll_no:
                errors.append(f"Row {index+2}: Student Name, Admission No, and Roll No are required. Values parsed: Name='{student_name}', Admission='{admission_no}', Roll='{roll_no}'")
                continue
                
            # Parse subjects with loose matching
            subjects_list = []
            standard_subjects_map = {
                "English": ["english"],
                "Hindi": ["hindi"],
                "Mathematics": ["mathematics", "math", "maths"],
                "Science": ["science"],
                "Social Science": ["social science", "social_science", "social", "sst"],
                "Computer": ["computer", "it", "information technology"]
            }
            
            for sub_display, kws in standard_subjects_map.items():
                obtained = None
                for kw in kws:
                    obtained = find_row_val(r, [f"{kw} obtained", f"{kw}_obtained", f"{kw} marks", f"{kw}_marks", f"{kw} score"])
                    if obtained is not None:
                        break
                if obtained is None:
                    obtained = find_row_val(r, kws)
                    
                if obtained is not None:
                    try:
                        obtained_val = float(obtained)
                        maximum_val = 100.0
                        sub_pct = (obtained_val / maximum_val * 100) if maximum_val > 0 else 0
                        grade = "A+" if sub_pct >= 90 else "A" if sub_pct >= 80 else "B" if sub_pct >= 70 else "C" if sub_pct >= 60 else "D" if sub_pct >= 50 else "E" if sub_pct >= 35 else "F"
                        
                        subjects_list.append({
                            "subject": sub_display,
                            "max_marks": maximum_val,
                            "obtained_marks": obtained_val,
                            "grade": grade,
                            "remarks": "Excellent" if sub_pct >= 75 else "Good" if sub_pct >= 50 else "Needs Improvement"
                        })
                    except (ValueError, TypeError):
                        pass
            
            # Check for any other column in row that is a subject but not in standard map or metadata
            metadata_keys = [
                "student", "name", "admission", "roll", "father", "class", "section", "year",
                "total", "percentage", "pct", "percent", "division", "grade", "status", "result",
                "remarks", "s.no", "sno", "s_no", "sl.no", "sl_no", "serial"
            ]
            for k, v in r.items():
                k_clean = str(k).lower().strip()
                if any(m in k_clean for m in metadata_keys):
                    continue
                is_standard = False
                for kws in standard_subjects_map.values():
                    if any(kw in k_clean for kw in kws):
                        is_standard = True
                        break
                if is_standard:
                    continue
                
                if v is not None and str(v).strip() != "":
                    try:
                        obtained_val = float(v)
                        maximum_val = 100.0
                        sub_pct = (obtained_val / maximum_val * 100) if maximum_val > 0 else 0
                        grade = "A+" if sub_pct >= 90 else "A" if sub_pct >= 80 else "B" if sub_pct >= 70 else "C" if sub_pct >= 60 else "D" if sub_pct >= 50 else "E" if sub_pct >= 35 else "F"
                        
                        subjects_list.append({
                            "subject": str(k).strip().title(),
                            "max_marks": maximum_val,
                            "obtained_marks": obtained_val,
                            "grade": grade,
                            "remarks": "Passed" if sub_pct >= 35 else "Failed"
                        })
                    except (ValueError, TypeError):
                        pass
                        
            if not subjects_list:
                # Consolidated results (no subject-wise marks)
                total_val = find_row_val(r, ["total marks", "total_marks", "total", "marks"])
                percentage_val = find_row_val(r, ["percentage", "pct", "percent"])
                
                if total_val is not None or percentage_val is not None:
                    try:
                        total = float(total_val) if total_val is not None else 0.0
                    except (ValueError, TypeError):
                        total = 0.0
                        
                    try:
                        if percentage_val is not None:
                            pct_str = str(percentage_val).replace("%", "").strip()
                            percentage = float(pct_str)
                        else:
                            percentage = 0.0
                    except (ValueError, TypeError):
                        percentage = 0.0
                        
                    division_val = find_row_val(r, ["division", "div"])
                    division = str(division_val).strip() if division_val is not None else ""
                    if not division:
                        if percentage >= 60.0:
                            division = "First"
                        elif percentage >= 45.0:
                            division = "Second"
                        elif percentage >= 35.0:
                            division = "Third"
                        else:
                            division = "Fail"
                            
                    grade_val = find_row_val(r, ["grade"])
                    grade = str(grade_val).strip() if grade_val is not None else ""
                    if not grade:
                        if percentage >= 90:
                            grade = "A+"
                        elif percentage >= 80:
                            grade = "A"
                        elif percentage >= 70:
                            grade = "B"
                        elif percentage >= 60:
                            grade = "C"
                        elif percentage >= 50:
                            grade = "D"
                        elif percentage >= 35:
                            grade = "E"
                        else:
                            grade = "F"
                            
                    status_val = find_row_val(r, ["status", "result"])
                    status = str(status_val).strip().title() if status_val is not None else ""
                    if not status:
                        status = "Pass" if percentage >= 35.0 else "Fail"
                        
                    metrics = {
                        "total": total,
                        "percentage": percentage,
                        "division": division,
                        "grade": grade,
                        "status": status
                    }
                else:
                    errors.append(f"Row {index+2} ({student_name}): No valid subject scores or consolidated marks found. Column keys in file: {list(r.keys())}")
                    continue
            else:
                metrics = calculate_result_metrics(subjects_list)
            
            doc = {
                "student_name": student_name,
                "admission_no": admission_no,
                "roll_no": roll_no,
                "father_name": father_name,
                "class": class_name,
                "section": section,
                "academic_year": academic_year,
                "subjects": subjects_list,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
            doc.update(metrics)
            
            await db.results.update_one(
                {"admission_no": admission_no, "academic_year": academic_year},
                {"$set": doc},
                upsert=True
            )
            imported += 1
        except Exception as ex:
            errors.append(f"Row {index+2}: Unexpected error: {str(ex)}")
            
    # recalculate_class_toppers is now handled dynamically in front-end class portals
            
    return {
        "message": f"Bulk import complete. Imported: {imported} records.",
        "failed_count": len(errors),
        "errors": errors[:50]
    }

@router.get("/results/export/excel/{class_name}")
async def export_results_excel(class_name: str, admin=Depends(get_current_admin)):
    cursor = db.results.find({"class": class_name})
    results = await cursor.to_list(1000)
    
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Class {class_name} Results"
        
        # Write headers
        headers = ["Student Name", "Admission No", "Roll No", "Father Name", "Class", "Section", "Academic Year", "Total Marks", "Percentage", "Division", "Grade", "Status"]
        ws.append(headers)
        
        for r in results:
            ws.append([
                r.get("student_name", ""),
                r.get("admission_no", ""),
                r.get("roll_no", ""),
                r.get("father_name", ""),
                r.get("class", ""),
                r.get("section", ""),
                r.get("academic_year", ""),
                r.get("total", 0.0),
                r.get("percentage", 0.0),
                r.get("division", ""),
                r.get("grade", ""),
                r.get("status", "")
            ])
            
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=class_{class_name}_results.xlsx"}
        )
    except Exception as e:
        # Fallback to CSV if openpyxl fails or raises exception
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Student Name", "Admission No", "Roll No", "Father Name", "Class", "Section", "Academic Year", "Total Marks", "Percentage", "Division", "Grade", "Status"])
        for r in results:
            writer.writerow([
                r.get("student_name", ""),
                r.get("admission_no", ""),
                r.get("roll_no", ""),
                r.get("father_name", ""),
                r.get("class", ""),
                r.get("section", ""),
                r.get("academic_year", ""),
                r.get("total", 0.0),
                r.get("percentage", 0.0),
                r.get("division", ""),
                r.get("grade", ""),
                r.get("status", "")
            ])
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=class_{class_name}_results.csv"}
        )

# =====================================================
# PDF EXPORTS (REPORTLAB)
# =====================================================

@router.get("/results/export/pdf/{class_name}")
async def export_results_pdf(class_name: str, admin=Depends(get_current_admin)):
    cursor = db.results.find({"class": class_name}).sort([("percentage", -1)])
    results = await cursor.to_list(1000)
    
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Title"],
            fontSize=16,
            textColor=colors.HexColor("#08295a"),
            spaceAfter=10
        )
        
        subtitle_style = ParagraphStyle(
            name="SubTitleStyle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#555555"),
            alignment=1, # Center
            spaceAfter=20
        )
        
        story.append(Paragraph("EKLAVYA MODEL RESIDENTIAL SCHOOL", title_style))
        story.append(Paragraph(f"Dornala, Prakasam, Andhra Pradesh — Class {class_name} Results Consolidated", subtitle_style))
        story.append(Spacer(1, 10))
        
        # Build Table data
        data = [["Rank", "Roll No", "Admission No", "Student Name", "Father Name", "Total", "%", "Grade", "Status"]]
        for idx, r in enumerate(results):
            data.append([
                str(idx + 1),
                r.get("roll_no", ""),
                r.get("admission_no", ""),
                r.get("student_name", ""),
                r.get("father_name", ""),
                str(r.get("total", "")),
                f"{r.get('percentage', 0.0)}%",
                r.get("grade", ""),
                r.get("status", "")
            ])
            
        t = Table(data, colWidths=[35, 50, 70, 130, 110, 45, 45, 35, 40])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#08295a")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#f6faff"), colors.white]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dce7f5")),
            ('TEXTCOLOR', (8,1), (8,-1), colors.HexColor("#198754")), # Pass text green
        ]))
        
        # Color fails red
        for idx, r in enumerate(results):
            if r.get("status") == "Fail":
                t.setStyle(TableStyle([
                    ('TEXTCOLOR', (8, idx + 1), (8, idx + 1), colors.HexColor("#dc3545")),
                ]))
                
        story.append(t)
        doc.build(story)
        pdf_buffer.seek(0)
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=class_{class_name}_results.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")

@router.get("/results/student/{admission_no}/pdf")
async def export_student_marksheet_pdf(admission_no: str):
    result = await db.results.find_one({"admission_no": admission_no})
    if not result:
        raise HTTPException(status_code=404, detail="Student result not found")
        
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.graphics.shapes import Drawing, Rect
        
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        story = []
        
        styles = getSampleStyleSheet()
        
        header_style = ParagraphStyle(
            name="HeaderTitle",
            parent=styles["Title"],
            fontSize=18,
            textColor=colors.HexColor("#08295a"),
            fontName="Helvetica-Bold",
            spaceAfter=2
        )
        
        sub_header_style = ParagraphStyle(
            name="SubHeader",
            parent=styles["Normal"],
            fontSize=9,
            alignment=1, # Center
            textColor=colors.HexColor("#333333"),
            spaceAfter=15
        )
        
        title_style = ParagraphStyle(
            name="DocTitle",
            parent=styles["Title"],
            fontSize=14,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1b2e4b"),
            spaceAfter=15
        )
        
        label_style = ParagraphStyle(
            name="Label",
            parent=styles["Normal"],
            fontSize=10,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#08295a")
        )
        
        val_style = ParagraphStyle(
            name="Value",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#333333")
        )
        
        story.append(Paragraph("EKLAVYA MODEL RESIDENTIAL SCHOOL", header_style))
        story.append(Paragraph("Dornala, Prakasam, Andhra Pradesh, Pin: 523315<br/>(An Autonomous Organisation under Ministry of Tribal Affairs, Govt of India)", sub_header_style))
        story.append(Paragraph("REPORT CARD / MARKSHEET", title_style))
        story.append(Spacer(1, 10))
        
        # Student Info Grid
        info_data = [
            [
                Paragraph("Student Name:", label_style), Paragraph(result.get("student_name", ""), val_style),
                Paragraph("Admission No:", label_style), Paragraph(result.get("admission_no", ""), val_style)
            ],
            [
                Paragraph("Father Name:", label_style), Paragraph(result.get("father_name", ""), val_style),
                Paragraph("Roll Number:", label_style), Paragraph(result.get("roll_no", ""), val_style)
            ],
            [
                Paragraph("Class:", label_style), Paragraph(f"Class {result.get('class', '')}", val_style),
                Paragraph("Section & Year:", label_style), Paragraph(f"{result.get('section', '')} | {result.get('academic_year', '')}", val_style)
            ]
        ]
        info_table = Table(info_data, colWidths=[100, 160, 100, 160])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fbfe")),
            ('PADDING', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dbe7f5")),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Subjects Scores Table
        sub_headers = ["Subject", "Maximum Marks", "Obtained Marks", "Grade", "Remarks"]
        scores_data = [sub_headers]
        
        for s in result.get("subjects", []):
            scores_data.append([
                s.get("subject", ""),
                str(s.get("max_marks", 100.0)),
                str(s.get("obtained_marks", 0.0)),
                s.get("grade", ""),
                s.get("remarks", "")
            ])
            
        # Add totals row
        scores_data.append([
            "Grand Total",
            str(sum(float(s.get("max_marks", 100)) for s in result.get("subjects", []))),
            str(result.get("total", 0.0)),
            result.get("grade", ""),
            result.get("status", "")
        ])
        
        scores_table = Table(scores_data, colWidths=[160, 95, 95, 65, 115])
        scores_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#08295a")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'), # left align subjects
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dce7f5")),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#eef3f9")), # Total background
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (4,-1), (4,-1), colors.HexColor("#198754") if result.get("status") == "Pass" else colors.HexColor("#dc3545")),
        ]))
        story.append(scores_table)
        story.append(Spacer(1, 20))
        
        # Summary metrics
        summary_data = [
            [
                Paragraph("Percentage:", label_style), Paragraph(f"{result.get('percentage', 0.0)}%", val_style),
                Paragraph("Overall Grade:", label_style), Paragraph(result.get("grade", ""), val_style),
                Paragraph("Division:", label_style), Paragraph(result.get("division", ""), val_style)
            ]
        ]
        summary_table = Table(summary_data, colWidths=[80, 90, 90, 90, 80, 100])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#e8f4fd")),
            ('PADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#bce1fd")),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # QR Code and Signature layout
        # We will embed a QR code from a public server API that links to the verification endpoint
        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=100x100&data=http://emrsdornala.edu.in/api/results/student/{admission_no}"
        
        # Fetch QR Code bytes defensively, fallback if offline
        qr_elem = None
        try:
            import urllib.request
            # Bounded fetch to avoid freezing
            req = urllib.request.Request(qr_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=3) as response:
                qr_bytes = response.read()
                qr_elem = Image(io.BytesIO(qr_bytes), width=80, height=80)
        except Exception as e:
            logger.warning(f"Failed to fetch QR Code image dynamically, using placeholder: {str(e)}")
            
        sign_label_style = ParagraphStyle(
            name="SignLabel",
            parent=styles["Normal"],
            fontSize=10,
            fontName="Helvetica-Bold",
            alignment=1, # Center
            textColor=colors.HexColor("#333333")
        )
        
        # If QR element is missing, make a simple placeholder image or empty table cell
        qr_cell = qr_elem if qr_elem else Paragraph("QR Verify", label_style)
        
        sign_data = [
            [
                qr_cell,
                Paragraph("__________________<br/><br/>Class Teacher", sign_label_style),
                Paragraph("__________________<br/><br/>Principal / Controller", sign_label_style)
            ]
        ]
        sign_table = Table(sign_data, colWidths=[150, 190, 190])
        sign_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(sign_table)
        
        doc.build(story)
        pdf_buffer.seek(0)
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=marksheet_{admission_no}.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate marksheet PDF: {str(e)}")
